# -*- coding: utf-8 -*-

from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect, reverse
from django.core.mail import send_mail
from django.contrib import messages
from django.conf import settings
from datetime import datetime
from openpyxl import load_workbook
from Main.models import Configure, TempEmployer, UpdateEmployer, Widget, TypeStatus, WidgetStatus, Employer, \
    StatusEmployer, Info, TypeViolations, TypeNotify, Notify
from Main.tools import get_profile
from Main.decorators import superuser_only
import os
import re

######################################################################################################################


@superuser_only
def configure_list(request) -> HttpResponse:
    """
    Отображение конфигураций для настроки и обслуживанию информационной системы
    :param request:
    :return:
    """
    context = {
        'current_profile': get_profile(user=request.user),
        'title': 'Список конфигураций',
        'list_configure': Configure.objects.all(),
    }
    return render(request=request, template_name='configure/configure_list.html', context=context, )


######################################################################################################################


@superuser_only
def employer_status_sync(request):
    """
    Синхронизация статуса карточек работодателей для перевода их на новую версию информационной системы.
    :param request:
    :return:
    TODO: Удалить после перехода на новую версию
    """
    list_employer = Employer.objects.all()
    count = 0
    for employer in list_employer:
        if not StatusEmployer.objects.filter(employer=employer).exists():
            old_status = TypeStatus.objects.filter(status=employer.Status).first()
            if old_status:
                StatusEmployer.objects.create(employer=employer, type_status=old_status)
                count = count + 1
    messages.info(
        request,
        f'Обновлено {count} статусов в карточках работодателей-нарушителей'
    )
    return redirect(reverse('configure_list'))


######################################################################################################################


@superuser_only
def employer_violations_sync(request):
    """
    Синхронизация информации о правонарушении для перевода их на новую версию информационной системы.
    :param request:
    :return:
    TODO: Удалить после перехода на новую версию
    """
    list_violations = Info.objects.filter(title=None)
    count = 0
    for violation in list_violations:
        if not violation.title:
            old_violation = TypeViolations.objects.filter(old_name=violation.Name).first()
            if old_violation:
                violation.title = old_violation
                violation.save(update_fields=['title'])
                count = count + 1
    messages.info(
        request,
        f'Обновлено {count} правонарушений в карточках работодателей-нарушителей'
    )
    return redirect(reverse('configure_list'))


######################################################################################################################


@superuser_only
def employer_notify_sync(request):
    """
    Синхронизация информации об информировании работодателя центром занятости для перевода их на новую версию
    информационной системы.
    :param request:
    :return:
    TODO: Удалить после перехода на новую версию
    """
    list_notify = Notify.objects.filter(title=None)
    count = 0
    for notify in list_notify:
        if not notify.title:
            old_notify = TypeNotify.objects.filter(old_name=notify.Method).first()
            if old_notify:
                notify.title = old_notify
                notify.save(update_fields=['title'])
                count = count + 1
    messages.info(
        request,
        f'Обновлено {count} информирований в карточках работодателей-нарушителей'
    )
    return redirect(reverse('configure_list'))


######################################################################################################################


@superuser_only
def employer_set_owner(request):
    """
    Назначение владельцев карточек (отделов) в зависимости от отдела пользователя, который создал карточку.
    :param request:
    :return:
    TODO: Удалить после перехода на новую версию
    """
    list_employer = Employer.objects.all()
    count = 0
    for employer in list_employer:
        department = employer.Owner.department
        if department:
            employer.owner_department = department
            employer.save(update_fields=['owner_department'])

    messages.info(
        request,
        f'Количество карточек, для которых назначены владельцы - {count}.'
    )

    return redirect(reverse('configure_list'))


######################################################################################################################


@superuser_only
def send_email(request):
    """
    Тестирование отправки сообщения на электронную почту
    :param request:
    :return:
    """
    if not request.user.is_superuser:
        return redirect(reverse('index'))

    email = ['postmaster@webmail.omskzan.ru']
    data = """
    Привет.
    Это тестовое письмо.
    ---
    С Уважением Робот.
    """
    send_mail('Заголовок', data, settings.DEFAULT_FROM_EMAIL, email, fail_silently=False)
    messages.info(
        request,
        f'Отправлено тестовое электронное письмо на адрес {email}.'
    )

    return redirect(reverse('configure_list'))


######################################################################################################################


@superuser_only
def employer_load(request):
    """
    Загрузка информации об организациях из подготовленного файла в Катарсисе
    :param request:
    :return:
    """

    if request.POST:

        def send_blank(value, decoding=False):
            if value:
                if decoding:
                    result = re.sub('_x000d_', ' ', value, flags=re.IGNORECASE)
                else:
                    result = value
            else:
                result = ''
            return result

        start_time = datetime.now()
        TempEmployer.objects.all().delete()
        for count, x in enumerate(request.FILES.getlist('files')):
            file = settings.UPLOAD_FILE + str(count) + '.xlsx'

            def process(f):
                with open(file, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)

            process(x)
            workbook = load_workbook(filename=file)
            sheet = workbook.active
            if sheet.max_row > 2:
                list_temp_employer = []
                for row in range(3, sheet.max_row + 1):
                    temp_employer = TempEmployer(
                        Number=send_blank(value=sheet.cell(row=row, column=1).value),
                        Title=send_blank(value=sheet.cell(row=row, column=2).value, decoding=True),
                        INN=send_blank(value=sheet.cell(row=row, column=3).value),
                        OGRN=send_blank(value=sheet.cell(row=row, column=4).value),
                        JurAddress=send_blank(value=sheet.cell(row=row, column=5).value),
                        FactAddress=send_blank(value=sheet.cell(row=row, column=6).value),
                        Contact=send_blank(value=sheet.cell(row=row, column=7).value),
                    )
                    event_date = send_blank(sheet.cell(row=row, column=8).value)
                    if isinstance(event_date, datetime):
                        temp_employer.EventDate = event_date.strftime("%Y-%m-%d")
                    list_temp_employer.append(temp_employer)
                TempEmployer.objects.bulk_create(list_temp_employer)

                messages.info(
                    request,
                    f'Файл {x.name} содержит записей организаций - {sheet.max_row - 2}.'
                )
            else:
                messages.error(
                    request,
                    f'Файл {x.name} не содержит данные.'
                )
            workbook.close()
            os.remove(file)
        update_employer = UpdateEmployer(
            count_employer=TempEmployer.objects.all().count(),
            time_spent=(datetime.now() - start_time).seconds,
        )
        update_employer.save()
        messages.success(
            request,
            'Успешно загружено {0} записей организаций из Катарсис за {1} секунд(ы).'.format(
                update_employer.count_employer,
                update_employer.time_spent,
            )
        )
    context = {
        'current_profile': get_profile(user=request.user),
        'title': 'Загрузить работодателей',
        'list_breadcrumb': ((reverse('configure_list'), 'Список конфигураций'),),
        'list_upload': UpdateEmployer.objects.all()[:10],
    }
    return render(request=request, template_name='configure/employer_load.html', context=context, )


######################################################################################################################


@superuser_only
def widget_list(request):
    """
    Настройка виджетов, отображение текущих фильтров и их изменение
    :param request:
    :return:
    """
    if request.GET:
        id_widget = int(request.GET.get('id', 0))
        widget = get_object_or_404(Widget, id=id_widget)
        list_filter = WidgetStatus.objects.filter(widget=widget)
        context = {
            'widget': widget,
            'list_filter': list_filter,
        }
        return render(request=request, template_name='configure/filter_modal.html', context=context, )
    else:
        if request.POST:
            id_widget = request.POST.get('id_widget', 0)
            selected_filter = request.POST.getlist('selected_filter')
            widget = get_object_or_404(Widget, id=id_widget)
            WidgetStatus.objects.filter(widget=widget).exclude(id__in=selected_filter).update(checked=False)
            WidgetStatus.objects.filter(id__in=selected_filter).update(checked=True)
            messages.info(
                request,
                'Обновление настроек виджета "{0}" завершено.'.format(widget.title)
            )
            return redirect(reverse('widget_list'))
        else:
            list_widget = []
            list_status = TypeStatus.objects.all()
            for widget in Widget.objects.all():
                for status in list_status:
                    widget_filter, created = WidgetStatus.objects.get_or_create(widget=widget, status=status)
                list_widget.append([widget, WidgetStatus.objects.filter(widget=widget)])
            context = {
                'current_profile': get_profile(user=request.user),
                'title': 'Список фильтров',
                'list_breadcrumb': ((reverse('configure_list'), 'Список конфигураций'),),
                'list_widget': list_widget,
            }
            return render(request=request, template_name='configure/filter_list.html', context=context, )


######################################################################################################################
